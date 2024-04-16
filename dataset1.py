# This is Data set of House Price


import os
import math
import numpy as np
import pandas as pd
import seaborn as sns
from datetime import datetime
from IPython.display import display
import openpyxl

from sklearn.decomposition import PCA
from sklearn.linear_model import Ridge
from sklearn.linear_model import Lasso
from sklearn.linear_model import ElasticNet
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import PolynomialFeatures
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error

import matplotlib.pyplot as plt
plt.rcParams['figure.figsize'] = [10, 6]

import warnings
warnings.filterwarnings('ignore')

# Read the CSV file
try:
    df = pd.read_csv("/Users/phonemyatmin/Downloads/Housing.csv")
    print("Data loaded successfully!")
except FileNotFoundError:
    print("File not found. Please check the file path.")

# Display a sample of the DataFrame
print(df.sample(10))

# Define the target variable
target = 'price'  # Assuming 'price' is the target variable

# Data Stats
data_stats = df.describe()

# Encode categorical columns
df_encoded = pd.get_dummies(df)


# Plot target variable distribution
plt.figure(figsize=[10, 6])
sns.distplot(df[target], color='red', hist_kws=dict(edgecolor="black", linewidth=2), bins=30)
plt.title('Target Variable Distribution - Median Value of Homes ($1Ms)')

# Save the seaborn plot as an image
target_plot_path = "/Users/phonemyatmin/Downloads/target_variable_distribution.png"
plt.savefig(target_plot_path)

# Visualizing the categorical features
cf = df.select_dtypes(include=['object']).columns.tolist()
print('\033[1mVisualising Categorical Features:'.center(100))

n = 3
plt.figure(figsize=[15, 3 * math.ceil(len(cf) / n)])

for i in range(len(cf)):
    if df[cf[i]].nunique() <= 8:
        plt.subplot(math.ceil(len(cf) / n), n, i + 1)
        sns.countplot(df[cf[i]])  # Change color to '#FF5733'
    else:
        plt.subplot(3, 1, 3)
        sns.countplot(df[cf[i]])  # Change color to '#FF5733'

plt.tight_layout()

# Save the categorical features visualization as an image
categorical_plot_path = "/Users/phonemyatmin/Downloads/categorical_features_distribution.png"
plt.savefig(categorical_plot_path)

# Visualizing the numeric features
print('\033[1mNumeric Features Distribution'.center(130))

nf = df.select_dtypes(include=[np.number]).columns.tolist()

n = 3
plt.figure(figsize=[15, 3 * math.ceil(len(nf) / n)])
for i in range(len(nf)):
    plt.subplot(math.ceil(len(nf) / 3), n, i + 1)
    sns.distplot(df[nf[i]], hist_kws=dict(edgecolor="black", linewidth=2), bins=10, color=list(np.random.randint(0, 255, size=3)/255))
plt.tight_layout()

# Save the numeric features distribution plot as an image
numeric_plot_path = "/Users/phonemyatmin/Downloads/numeric_features_distribution.png"
plt.savefig(numeric_plot_path)

# Understanding the relationship between all the features
plt.figure(figsize=[20, 15])
plt.title('Pairplots for all the Feature')
g = sns.pairplot(df)
g.map_upper(sns.kdeplot, levels=4, color=".2")

# Save the pairplot as an image
pairplot_path = "/Users/phonemyatmin/Downloads/pairplot.png"
plt.savefig(pairplot_path)

# Define the path to the output Excel file
output_excel_path = "/Users/phonemyatmin/Downloads/Housing_output.xlsx"

# Create an ExcelWriter object to write to the Excel file
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    # Write the DataFrame to a sheet named 'Data Information'
    df.to_excel(writer, sheet_name='Data Information', index=False)

    # Write the data stats to a new sheet named 'Data Stats'
    data_stats.to_excel(writer, sheet_name='Data Stats')

    # Add the seaborn plot as an image to a new sheet
    sns_plot_sheet_name = "Target Variable Distribution"
    sns_plot_worksheet = writer.book.create_sheet(sns_plot_sheet_name)
    img = openpyxl.drawing.image.Image(target_plot_path)
    sns_plot_worksheet.add_image(img, 'B3')

    # Add the categorical features distribution plot as an image to a new sheet
    categorical_plot_sheet_name = "Categorical Features Distribution"
    categorical_plot_worksheet = writer.book.create_sheet(categorical_plot_sheet_name)
    img = openpyxl.drawing.image.Image(categorical_plot_path)
    categorical_plot_worksheet.add_image(img, 'B3')

    # Add the numeric features distribution plot as an image to a new sheet
    numeric_plot_sheet_name = "Numeric Features Distribution"
    numeric_plot_worksheet = writer.book.create_sheet(numeric_plot_sheet_name)
    img = openpyxl.drawing.image.Image(numeric_plot_path)
    numeric_plot_worksheet.add_image(img, 'B3')

    # Add the pairplot as an image to a new sheet
    pairplot_sheet_name = "Pairplot"
    pairplot_worksheet = writer.book.create_sheet(pairplot_sheet_name)
    img = openpyxl.drawing.image.Image(pairplot_path)
    pairplot_worksheet.add_image(img, 'B3')

print(f"DataFrame, Data Stats, Data Correlation, and Target Variable Distribution saved to Excel file: {output_excel_path}")



